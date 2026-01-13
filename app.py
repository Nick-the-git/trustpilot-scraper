import streamlit as st
import requests
from bs4 import BeautifulSoup
import io
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook

st.title("Trustpilot Review Scraper")
st.write("Extract reviews from any Trustpilot page and receive them via email.")

url = st.text_input("Trustpilot URL", placeholder="https://www.trustpilot.com/review/example.com")
num_pages = st.number_input("Number of pages to scrape", min_value=1, max_value=100, value=5)
email = st.text_input("Your email", placeholder="you@example.com")

def scrape_trustpilot_reviews(url, num_pages, min_length=20):
    reviews = []
    headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15'}
    
    progress = st.progress(0)
    status = st.empty()
    
    for page in range(1, num_pages + 1):
        status.text(f"Scraping page {page} of {num_pages}...")
        progress.progress(page / num_pages)
        
        paginated_url = f"{url}?page={page}"
        response = requests.get(paginated_url, headers=headers)
        
        if response.status_code != 200:
            continue
        
        html_content = response.content.decode('utf-8')
        soup = BeautifulSoup(html_content, 'html.parser')
        review_blocks = soup.find_all('article')
        
        for block in review_blocks:
            text_tag = block.find('p', {'data-service-review-text-typography': 'true'})
            date_tag = block.find('time')
            rating_tag = block.find('img', alt=lambda x: x and 'Rated' in x)
            
            review_text = text_tag.get_text(separator=' ', strip=True) if text_tag else ''
            review_date = date_tag['datetime'] if date_tag and 'datetime' in date_tag.attrs else 'Unknown'
            star_rating = rating_tag['alt'] if rating_tag else 'Not found'
            
            if len(review_text) >= min_length:
                reviews.append({
                    'text': review_text,
                    'rating': star_rating,
                    'date': review_date
                })
    
    status.text(f"Done! Found {len(reviews)} reviews.")
    return reviews

def create_xlsx(reviews):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reviews"
    
    ws['A1'] = 'Review Text'
    ws['B1'] = 'Rating'
    ws['C1'] = 'Date'
    
    ws.column_dimensions['A'].width = 100
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    
    for i, review in enumerate(reviews, start=2):
        ws[f'A{i}'] = review['text']
        ws[f'B{i}'] = review['rating']
        ws[f'C{i}'] = review['date']
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def send_email_with_xlsx(to_email, xlsx_content, company_name, review_count):
    gmail_address = st.secrets["GMAIL_ADDRESS"]
    gmail_password = st.secrets["GMAIL_APP_PASSWORD"]
    
    analysis_prompt = """Prompt:

Analyse the attached Trustpilot reviews for the impact of the product on end users.

Impact is defined as a change in an outcome caused by an organization. An impact can be positive or negative, intended or unintended. An outcome is the level of well-being experienced by an individual or group of people, or the condition of the natural environment.

We are particularly interested in the 'what', 'who', 'depth' dimensions of impact. Read more about these:

1. Detailed information on the 'what' here: https://impactfrontiers.org/norms/five-dimensions-of-impact/what/
2. Detailed information on the 'who' here: https://impactfrontiers.org/norms/five-dimensions-of-impact/who/
3. Detailed information on the 'how much' here, use this to focus on the 'depth' sub-dimension: https://impactfrontiers.org/norms/five-dimensions-of-impact/how-much/
4. Detailed information on 'contribution' here: https://impactfrontiers.org/norms/five-dimensions-of-impact/enterprise-contribution/
5. Detailed information on impact risks: https://impactfrontiers.org/norms/five-dimensions-of-impact/impact-risk/

Structure the output by sub-dimension and give me your analysis of each area (and how you arrived at this conclusion). Give me a confidence rating from 1-100% and your rationale why. Provide a count of how many of the reviews you analysed support the claim you're making. Then give me the percentage from the total reviews. Include every quote that supports the claim. This is for each sub-section of output you produce.

Separately, give me the number of reviews where you cannot find any examples of any impact occurring (i.e. the absence of any outcome change that relates to the outcome). Then give me this as a percentage of the total reviews. Give me a confidence rating from 1-100% and your rationale why.

Create a frequency distribution of the rating scores. The x-axis is score and should be 1,2,3,4,5 to represent each possible score on Trustpilot. The y-axis would be the number of reviews that have that score.

Create a second frequency distribution on 'impact outcome occurrence'. The x-axis should be different outcome types and the y-axis should include how many of the reviews demonstrate that impact occurring. Include a negative y-axis to count any reviews that mention negative change against that outcome. Give me a table that's structured by the outcome types identified above and, in the cells, include a quote from every review that demonstrates the outcome type.

Focus on accuracy and quality:
• Throughout this exercise, only give me information you are sure is accurate.
• Do not synthesize any reviews or numbers. Rely exclusively on information presented in the document.
• If you are uncertain about any information, clearly state the uncertainty and do not present it as fact.
• Triple check your work.
• Try and find instances that disprove or oppose your view (i.e. the reverse thesis). List what they are and your rationale for not aligning your view with that evidence."""

    html_content = f"""<!DOCTYPE html>
<html>
<head>
    <style>
        body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
        .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
        h1 {{ color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }}
        h2 {{ color: #34495e; margin-top: 30px; }}
        .stats {{ background: #f8f9fa; padding: 15px; border-radius: 8px; margin: 20px 0; }}
        .prompt-box {{ background: #fff; border: 1px solid #ddd; border-radius: 8px; padding: 20px; margin-top: 20px; }}
        .prompt-text {{ background: #f5f5f5; padding: 15px; border-radius: 5px; white-space: pre-wrap; font-size: 14px; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>Trustpilot Reviews Export</h1>
        
        <div class="stats">
            <strong>Company:</strong> {company_name}<br>
            <strong>Total reviews:</strong> {review_count}
        </div>
        
        <p>Your requested export is attached as an Excel file.</p>
        
        <div class="prompt-box">
            <h2>Analysis Prompt</h2>
            <p>Upload the attached Excel file to your favourite LLM (Claude, ChatGPT, Copilot etc.) along with a prompt to start your analysis. If this is for understanding impact you can use the prompt below as a starting point (feel free to edit). If you have other use cases you regularly use, please let the AI team know and we can include those too. Here's the prompt:</p>
            <div class="prompt-text">{analysis_prompt}</div>
        </div>
    </div>
</body>
</html>"""
    
    msg = MIMEMultipart()
    msg['From'] = gmail_address
    msg['To'] = to_email
    msg['Subject'] = f"Trustpilot Reviews - {company_name}"
    
    msg.attach(MIMEText(html_content, 'html'))
    
    attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    attachment.set_payload(xlsx_content)
    encoders.encode_base64(attachment)
    attachment.add_header('Content-Disposition', f'attachment; filename="trustpilot_reviews_{company_name}.xlsx"')
    msg.attach(attachment)
    
    with smtplib.SMTP('smtp.gmail.com', 587) as server:
        server.starttls()
        server.login(gmail_address, gmail_password)
        server.send_message(msg)

if st.button("Scrape and Email Reviews"):
    if not url:
        st.error("Please enter a Trustpilot URL")
    elif "trustpilot.com/review/" not in url:
        st.error("Please enter a valid Trustpilot URL")
    elif not email:
        st.error("Please enter your email address")
    else:
        reviews = scrape_trustpilot_reviews(url, num_pages)
        
        if reviews:
            xlsx_content = create_xlsx(reviews)
            company_name = url.replace("https://www.trustpilot.com/review/", "").split("?")[0]
            
            try:
                send_email_with_xlsx(email, xlsx_content, company_name, len(reviews))
                st.success(f"Done! {len(reviews)} reviews sent to {email}")
            except Exception as e:
                st.error(f"Failed to send email: {type(e).__name__}: {e}")
                import traceback
                st.code(traceback.format_exc())
            
            st.download_button(
                label="Download Excel",
                data=xlsx_content,
                file_name=f"trustpilot_reviews_{company_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("No reviews found. Check the URL and try again.")
